import socket
import tkinter as tk
from tkinter import messagebox
import requests
import folium
import os
import webbrowser
from ipwhois import IPWhois
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os.path

# Function to get ISP and ASN information
def get_isp_and_asn(ip):
    try:
        obj = IPWhois(ip)
        result = obj.lookup_whois()
        isp = result.get('asn_description', 'Unknown ISP')
        asn = result.get('asn', 'Unknown ASN')
        return isp, asn
    except Exception as e:
        return "Error retrieving ISP/ASN", str(e)

# Function to get the IP location (existing function)
def get_ip_location(ip):
    try:
        response = requests.get(f'http://ip-api.com/json/{ip}')
        return response.json()
    except Exception as e:
        messagebox.showerror("Error", f"Could not retrieve data: {e}")
        return None

# Function to scan ports (basic scan for common ports)
def scan_ports(ip, ports=[80, 443, 22, 21, 25, 8080]):
    open_ports = []
    for port in ports:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((ip, port))  # 0 means open
        if result == 0:
            open_ports.append(port)
        sock.close()
    return open_ports

# Function to append data to Excel file
def append_to_excel(ip, location_name, isp, asn, open_ports):
    excel_file = 'ip_data.xlsx'

    # Get the current date and time
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")  # Format: Year-Month-Day Hour:Minute:Second

    # Check if the file already exists
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add header row if creating a new file
        sheet.append(["IP Address", "Location", "ISP", "ASN", "Open Ports", "Date & Time"])

    # Append new data with timestamp and open ports
    sheet.append([ip, location_name, isp, asn, ", ".join(map(str, open_ports)), timestamp])

    # Save the workbook
    workbook.save(excel_file)
    messagebox.showinfo("Excel Update", f"Data added to {excel_file}")

# Function to show map, perform port scan, and display ISP/ASN information
def show_map():
    ip = entry.get()
    location_data = get_ip_location(ip)
    
    # Get ISP and ASN details
    isp, asn = get_isp_and_asn(ip)

    if location_data and location_data['status'] == 'success':
        lat = location_data['lat']
        lon = location_data['lon']
        location_name = location_data['city'] + ', ' + location_data['country']

        # Perform a port scan
        open_ports = scan_ports(ip)

        # Display ISP, ASN, and port scan information
        messagebox.showinfo("IP Information", f"IP: {ip}\nLocation: {location_name}\nISP: {isp}\nASN: {asn}\nOpen Ports: {', '.join(map(str, open_ports))}")

        # Append results to Excel spreadsheet with timestamp and open ports
        append_to_excel(ip, location_name, isp, asn, open_ports)

        # Create a map centered around the IP's location
        map_ = folium.Map(location=[lat, lon], zoom_start=10)
        folium.Marker([lat, lon], tooltip=location_name).add_to(map_)
        
        # Save the map to an HTML file
        map_file = 'ip_location_map.html'
        map_.save(map_file)

        # Open the map in the default web browser
        webbrowser.open(f'file://{os.path.realpath(map_file)}')
    else:
        messagebox.showerror("Error", "Invalid IP address or unable to retrieve location.")

# Create the main window
root = tk.Tk()
root.title("IP Address Locator")

# Create and place the entry widget
label = tk.Label(root, text="Enter IP Address:")
label.pack(pady=10)

entry = tk.Entry(root, width=30)
entry.pack(pady=10)

# Create and place the button
button = tk.Button(root, text="Show Location on Map", command=show_map)
button.pack(pady=20)

# Run the application
root.mainloop()